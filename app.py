import streamlit as st
import pandas as pd
import difflib
import re
import io
import os
import json
import base64
from io import BytesIO
from collections import Counter

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

st.set_page_config(page_title="GridSync", page_icon="📊", layout="wide")

# =====================================================================
# Shared utilities
# =====================================================================

def normalize(value):
    s = "" if value is None else str(value)
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def similarity(a, b):
    na, nb = normalize(a), normalize(b)
    if not na and not nb:
        return 1.0
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    return difflib.SequenceMatcher(None, na, nb).ratio()


def read_uploaded_file(uploaded_file):
    ext = os.path.splitext(uploaded_file.name)[1].lower()
    if ext == ".csv":
        return pd.read_csv(uploaded_file, dtype=str, keep_default_na=False)
    return pd.read_excel(uploaded_file, dtype=str, keep_default_na=False)


def get_file_bytes(uploaded_file):
    uploaded_file.seek(0)
    data = uploaded_file.read()
    uploaded_file.seek(0)
    return data


def read_df_from_bytes(file_bytes, filename):
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".csv":
        return pd.read_csv(BytesIO(file_bytes), dtype=str, keep_default_na=False)
    return pd.read_excel(BytesIO(file_bytes), dtype=str, keep_default_na=False)


def to_excel_bytes(sheets: dict):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, df in sheets.items():
            safe_name = str(name)[:31] if name else "Sheet1"
            if df is None or df.empty:
                df = pd.DataFrame([{"Info": "No rows found"}])
            df.to_excel(writer, sheet_name=safe_name, index=False)
    return output.getvalue()


def combine_cols(row, cols):
    parts = [str(row[c]) for c in cols if c in row and pd.notna(row[c]) and str(row[c]).strip()]
    return " ".join(parts)


def first_nonempty(row, cols):
    for c in cols:
        if c in row and pd.notna(row[c]) and str(row[c]).strip():
            return row[c]
    return ""


EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
IMG_PATTERN = re.compile(r"image|packshot|pack ?shot|photo|img|picture|visual", re.I)
CGC_PATTERN = re.compile(r"categ|group|class", re.I)

NAME_PRIORITY_PATTERNS = [
    re.compile(r"group.?name", re.I),
    re.compile(r"class.?name", re.I),
    re.compile(r"sku.?name", re.I),
    re.compile(r"product.?name", re.I),
    re.compile(r"item.?name", re.I),
    re.compile(r"\bgroup\b", re.I),
    re.compile(r"\bclass\b", re.I),
    re.compile(r"\bsku\b", re.I),
    re.compile(r"\bitem\b", re.I),
]

ATTR_STORE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "gridsync_attributes.json")

# ---------------------------------------------------------------------
# Column / image detection (used by Merge, Compare, Catalog, SKU Match)
# ---------------------------------------------------------------------

def guess_name_columns(df, exclude_cols=None):
    """Detects the most likely product/SKU/group name column(s) — first by header
    keyword, then (if nothing matches) by text length and uniqueness, so it works
    even when the file uses completely custom headers."""
    exclude_cols = exclude_cols or set()
    cols = [c for c in df.columns if c not in exclude_cols]

    for pattern in NAME_PRIORITY_PATTERNS:
        hits = [c for c in cols if pattern.search(c)]
        if hits:
            return hits[:2]

    best_col, best_score = None, -1.0
    for c in cols:
        vals = df[c].astype(str).str.strip()
        vals = vals[vals != ""]
        if len(vals) == 0:
            continue
        avg_len = vals.str.len().mean()
        uniqueness = vals.nunique() / max(len(vals), 1)
        score = avg_len * uniqueness
        if score > best_score:
            best_score, best_col = score, c
    return [best_col] if best_col else []


def guess_image_columns(df, embedded_cols_set):
    keyword_cols = [c for c in df.columns if IMG_PATTERN.search(c)]
    url_cols = []
    for c in df.columns:
        sample = df[c].astype(str).str.strip().head(30)
        if sample.str.lower().str.startswith("http").any():
            url_cols.append(c)
    combined = list(dict.fromkeys(list(embedded_cols_set) + keyword_cols + url_cols))
    return combined


def cluster_columns(column_lists, threshold=0.6):
    """Groups similar column headers across multiple files into unified fields,
    so files with different formats can still be merged automatically."""
    all_cols = []
    for cols in column_lists:
        all_cols.extend(cols)
    clusters = []
    for col in all_cols:
        placed = False
        for cluster in clusters:
            if similarity(col, cluster[0]) >= threshold:
                cluster.append(col)
                placed = True
                break
        if not placed:
            clusters.append([col])
    representatives = []
    for cluster in clusters:
        rep = Counter(cluster).most_common(1)[0][0]
        representatives.append(rep)
    return representatives


# ---------------------------------------------------------------------
# Embedded image handling (System SKU Catalog / SKU Match)
# ---------------------------------------------------------------------

def extract_embedded_images(file_bytes, filename):
    """Extracts images embedded directly in xlsx cells and maps them to
    {(data_row_index, column_name): image_bytes}. Not supported for .xls/.csv."""
    ext = os.path.splitext(filename)[1].lower()
    if ext != ".xlsx":
        return {}
    try:
        wb = load_workbook(BytesIO(file_bytes))
        ws = wb.active
        header_row = 1
        col_name_by_idx = {}
        for cell in ws[header_row]:
            if cell.value is not None:
                col_name_by_idx[cell.column] = str(cell.value).strip()

        images_map = {}
        for img in getattr(ws, "_images", []):
            try:
                anchor = img.anchor
                from_marker = anchor._from
                col_1idx = from_marker.col + 1
                row_0idx = from_marker.row
                col_name = col_name_by_idx.get(col_1idx)
                if not col_name:
                    continue
                data_row_idx = row_0idx - header_row
                if data_row_idx < 0:
                    continue
                images_map[(data_row_idx, col_name)] = img._data()
            except Exception:
                continue
        return images_map
    except Exception:
        return {}


def image_bytes_to_data_uri(img_bytes):
    try:
        pil_img = PILImage.open(BytesIO(img_bytes))
        fmt = (pil_img.format or "PNG").lower()
        mime = f"image/{fmt}"
    except Exception:
        mime = "image/png"
    b64 = base64.b64encode(img_bytes).decode("utf-8")
    return f"data:{mime};base64,{b64}"


def get_packshot_value(row, row_idx, candidate_cols, embedded_map):
    for c in candidate_cols:
        key = (row_idx, c)
        if key in embedded_map:
            return {"display": image_bytes_to_data_uri(embedded_map[key]), "bytes": embedded_map[key]}
    val = first_nonempty(row, candidate_cols)
    if val:
        return {"display": str(val), "bytes": None}
    return {"display": "", "bytes": None}


def build_match_excel(results_df, image_bytes_map):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        results_df.to_excel(writer, sheet_name="SKU Match", index=False)
        ws = writer.sheets["SKU Match"]
        headers = list(results_df.columns)
        col_letter = {h: get_column_letter(i + 1) for i, h in enumerate(headers)}

        for h in ["Client Packshot", "Matched Packshot"]:
            if h in col_letter:
                ws.column_dimensions[col_letter[h]].width = 16

        for row_idx in range(len(results_df)):
            excel_row = row_idx + 2
            has_image = row_idx in image_bytes_map.get("client", {}) or row_idx in image_bytes_map.get("matched", {})
            if has_image:
                ws.row_dimensions[excel_row].height = 60
            for key, col_name in [("client", "Client Packshot"), ("matched", "Matched Packshot")]:
                bytes_map = image_bytes_map.get(key, {})
                if row_idx in bytes_map and col_name in col_letter:
                    try:
                        xl_img = XLImage(io.BytesIO(bytes_map[row_idx]))
                        xl_img.width, xl_img.height = 70, 70
                        ws.add_image(xl_img, f"{col_letter[col_name]}{excel_row}")
                    except Exception:
                        pass
    return output.getvalue()


# ---------------------------------------------------------------------
# Data quality analysis (Merge & Format)
# ---------------------------------------------------------------------

SPECIAL_CHAR_PATTERN = re.compile(r"[^A-Za-z0-9\s\-\.,&()/]")


def analyze_data_quality(df, check_cols):
    """Scans the merged data for duplicates, leading/trailing spaces, and
    special characters, and returns a summary plus two detail tables."""
    issues = []
    for idx, row in df.iterrows():
        for col in check_cols:
            val = row[col]
            if val is None:
                continue
            s = str(val)
            if s == "":
                continue
            if s != s.strip():
                issues.append({"Row": idx + 2, "Column": col, "Issue": "Leading/Trailing Space", "Value": s})
            specials = sorted(set(SPECIAL_CHAR_PATTERN.findall(s)))
            if specials:
                issues.append({"Row": idx + 2, "Column": col, "Issue": "Special Character(s): " + " ".join(specials), "Value": s})
    issues_df = pd.DataFrame(issues)

    if check_cols:
        normalized_view = df[check_cols].apply(lambda col: col.astype(str).str.strip().str.lower())
        dup_mask = normalized_view.duplicated(keep=False)
    else:
        dup_mask = pd.Series([False] * len(df))
    duplicates_df = df[dup_mask].copy()
    if not duplicates_df.empty:
        duplicates_df.insert(0, "Row", duplicates_df.index + 2)

    space_count = len(issues_df[issues_df["Issue"] == "Leading/Trailing Space"]) if not issues_df.empty else 0
    special_count = len(issues_df[issues_df["Issue"].str.startswith("Special", na=False)]) if not issues_df.empty else 0

    summary_df = pd.DataFrame([
        {"Metric": "Total Rows", "Count": len(df)},
        {"Metric": "Duplicate Rows", "Count": int(dup_mask.sum())},
        {"Metric": "Cells with Leading/Trailing Space", "Count": space_count},
        {"Metric": "Cells with Special Characters", "Count": special_count},
    ])
    return summary_df, duplicates_df, issues_df


# ---------------------------------------------------------------------
# Attribute completeness check (Compare Files)
# ---------------------------------------------------------------------

def load_attribute_options():
    if os.path.exists(ATTR_STORE_PATH):
        try:
            with open(ATTR_STORE_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []


def save_attribute_options(options):
    try:
        with open(ATTR_STORE_PATH, "w", encoding="utf-8") as f:
            json.dump(sorted(set(options)), f, ensure_ascii=False, indent=2)
    except Exception:
        pass


ATTRIBUTE_NAME_PATTERN = re.compile(r"attribute.?name", re.I)
ATTRIBUTE_VALUE_PATTERN = re.compile(r"attribute.?value", re.I)


def guess_attribute_columns(df):
    """Detects the 'attribute_name' and 'attribute_value' columns in a long-format
    CGC attribute file (one row per attribute, per Category/Group/Class)."""
    name_col = next((c for c in df.columns if ATTRIBUTE_NAME_PATTERN.search(c)), None)
    value_col = next((c for c in df.columns if ATTRIBUTE_VALUE_PATTERN.search(c)), None)
    return name_col, value_col


def check_attribute_completeness_long(df, id_cols, attr_name_col, attr_value_col, attributes):
    """
    Works with long-format attribute files:
    category_name, group_name, class_name, attribute_name, attribute_value, ...
    Each Category/Group/Class combination spans multiple rows (one per attribute).
    For each selected attribute, finds every CGC combination where that attribute
    is either never tagged, or tagged with an empty value.
    """
    work = df.copy()
    work["_cgc_key"] = work[id_cols].astype(str).agg(" | ".join, axis=1)
    all_keys_df = work[["_cgc_key"] + id_cols].drop_duplicates(subset="_cgc_key")
    total_cgc = len(all_keys_df)

    summary_rows, missing_records = [], []
    for attr in attributes:
        attr_norm = normalize(attr)
        attr_rows = work[work[attr_name_col].astype(str).apply(normalize) == attr_norm]
        present_keys = set(attr_rows[attr_rows[attr_value_col].astype(str).str.strip() != ""]["_cgc_key"])
        missing_keys_df = all_keys_df[~all_keys_df["_cgc_key"].isin(present_keys)]
        missing_count = len(missing_keys_df)

        summary_rows.append({
            "Attribute": attr,
            "Found in File": "Yes" if not attr_rows.empty else "No",
            "Total CGC": total_cgc,
            "Filled": total_cgc - missing_count,
            "Missing": missing_count,
        })
        for _, r in missing_keys_df.iterrows():
            rec = {"CGC": " > ".join(str(r[c]) for c in id_cols)}
            rec.update({c: r[c] for c in id_cols})
            rec["Missing Attribute"] = attr
            missing_records.append(rec)

    return pd.DataFrame(summary_rows), pd.DataFrame(missing_records)


# =====================================================================
# Sidebar navigation
# =====================================================================

st.sidebar.title("📊 GridSync")
st.sidebar.caption("Automated data tools — merge, compare, and match with minimal manual work.")
tool = st.sidebar.radio(
    "Select a tool",
    ["Merge & Format", "Compare Files", "System SKU Catalog", "SKU Match"],
)

# =====================================================================
# TOOL 1: Merge & Format
# =====================================================================

if tool == "Merge & Format":
    st.header("Merge & Format")
    st.caption(
        "Upload files in any format. The tool automatically aligns matching columns across "
        "all files, merges them into one dataset, and produces a data quality report."
    )

    uploaded_files = st.file_uploader(
        "Upload files to merge", type=["xlsx", "xls", "csv"], accept_multiple_files=True, key="merge_files"
    )

    if uploaded_files:
        parsed = [(uf.name, read_uploaded_file(uf)) for uf in uploaded_files]
        source_signature = tuple(uf.name for uf in uploaded_files)

        if st.session_state.get("std_fields_source") != source_signature:
            st.session_state.std_fields = cluster_columns([df.columns.tolist() for _, df in parsed])
            st.session_state.std_fields_source = source_signature

        st.success(f"🤖 Auto-detected {len(st.session_state.std_fields)} common field(s) across {len(parsed)} file(s).")

        with st.expander("⚙️ Advanced: edit the standard field list"):
            c1, c2 = st.columns([4, 1])
            new_field = c1.text_input("Add a field", key="new_field_input", label_visibility="collapsed", placeholder="e.g. Client Name")
            if c2.button("+ Add Field") and new_field.strip():
                if new_field.strip() not in st.session_state.std_fields:
                    st.session_state.std_fields.append(new_field.strip())
                st.rerun()
            if st.session_state.std_fields:
                chip_cols = st.columns(min(len(st.session_state.std_fields), 6))
                for i, f in enumerate(list(st.session_state.std_fields)):
                    if chip_cols[i % len(chip_cols)].button(f"✕ {f}", key=f"remove_field_{f}"):
                        st.session_state.std_fields.remove(f)
                        st.rerun()

        std_fields = st.session_state.std_fields

        st.subheader("Column mapping")
        st.caption("Mapping is auto-detected. Open a file below only if you need to correct it.")
        mappings = {}
        for fname, df in parsed:
            with st.expander(f"📄 {fname} ({len(df)} rows)", expanded=False):
                file_map = {}
                map_cols = st.columns(2)
                for i, field in enumerate(std_fields):
                    best_col, best_score = None, 0.0
                    for c in df.columns:
                        sc = similarity(field, c)
                        if sc > best_score:
                            best_score, best_col = sc, c
                    options = ["— leave empty —"] + list(df.columns)
                    default_idx = options.index(best_col) if best_score >= 0.45 and best_col else 0
                    with map_cols[i % 2]:
                        choice = st.selectbox(field, options, index=default_idx, key=f"map_{fname}_{field}")
                    file_map[field] = None if choice == "— leave empty —" else choice
            mappings[fname] = (df, file_map)

        if st.button("Merge Files", type="primary"):
            all_rows = []
            for fname, (df, file_map) in mappings.items():
                for _, row in df.iterrows():
                    out = {field: (row[col] if col else "") for field, col in file_map.items()}
                    out["Source File"] = fname
                    all_rows.append(out)
            merged_df = pd.DataFrame(all_rows, columns=std_fields + ["Source File"])
            st.session_state.merged_df = merged_df

    if "merged_df" in st.session_state:
        merged_df = st.session_state.merged_df
        std_fields = st.session_state.std_fields

        st.success(f"✅ Merged output: {len(merged_df)} rows from {merged_df['Source File'].nunique()} file(s)")

        st.subheader("Rows per source file")
        st.bar_chart(merged_df["Source File"].value_counts())

        st.subheader("Filter data")
        st.caption(
            "Choose up to 5 columns and specific values — this filter applies to the data preview, "
            "the Data Quality Report, and the downloadable files below."
        )
        filter_cols = st.multiselect("Choose up to 5 columns to filter by", options=std_fields, key="merge_filter_cols")
        if len(filter_cols) > 5:
            st.warning("Only the first 5 selected columns will be used.")
            filter_cols = filter_cols[:5]

        filtered_df = merged_df.copy()
        for fc in filter_cols:
            options = sorted([v for v in merged_df[fc].dropna().unique().tolist() if str(v).strip() != ""])
            chosen = st.multiselect(f"{fc} values", options, key=f"filter_val_{fc}")
            if chosen:
                filtered_df = filtered_df[filtered_df[fc].isin(chosen)]

        is_filtered = bool(filter_cols) and any(
            st.session_state.get(f"filter_val_{fc}") for fc in filter_cols
        )

        st.dataframe(filtered_df.head(100), use_container_width=True)
        if len(filtered_df) > 100:
            st.caption(f"Showing first 100 of {len(filtered_df)} filtered rows.")

        summary_df, duplicates_df, issues_df = analyze_data_quality(filtered_df, std_fields)

        st.subheader("Data Quality Summary" + (" (filtered)" if is_filtered else ""))
        st.dataframe(summary_df, use_container_width=True, hide_index=True)

        d1, d2, d3 = st.columns(3)
        with d1:
            st.download_button(
                "⬇ Download Merged Data",
                data=to_excel_bytes({"Merged Data": merged_df}),
                file_name="merged-data.xlsx",
                mime=EXCEL_MIME,
            )
        with d2:
            st.download_button(
                "⬇ Download Data Quality Report" + (" (filtered)" if is_filtered else ""),
                data=to_excel_bytes({
                    "Summary": summary_df,
                    "Duplicate Rows": duplicates_df,
                    "Issues (Spaces & Special Characters)": issues_df,
                }),
                file_name="data-quality-report.xlsx",
                mime=EXCEL_MIME,
            )
        with d3:
            if filter_cols:
                st.download_button(
                    "⬇ Download Filtered Data",
                    data=to_excel_bytes({"Filtered Data": filtered_df}),
                    file_name="filtered-data.xlsx",
                    mime=EXCEL_MIME,
                )

# =====================================================================
# TOOL 2: Compare Files  (+ Attribute Completeness Check)
# =====================================================================

elif tool == "Compare Files":
    st.header("Compare Files")
    st.caption("Upload two files in any format — the key column is detected automatically.")

    c1, c2 = st.columns(2)
    df_a = df_b = None
    key_a = key_b = None

    with c1:
        file_a = st.file_uploader("POC Project File", type=["xlsx", "xls", "csv"], key="file_a")
        if file_a:
            df_a = read_uploaded_file(file_a)
            auto_key_a = guess_name_columns(df_a)
            default_idx_a = list(df_a.columns).index(auto_key_a[0]) if auto_key_a else 0
            st.caption(f"🤖 Auto-detected key column: **{auto_key_a[0] if auto_key_a else 'not found'}**")
            with st.expander("⚙️ Change key column"):
                key_a = st.selectbox("Key column", df_a.columns, index=default_idx_a, key="key_a_select")

    with c2:
        file_b = st.file_uploader("Pilot Project File", type=["xlsx", "xls", "csv"], key="file_b")
        if file_b:
            df_b = read_uploaded_file(file_b)
            auto_key_b = guess_name_columns(df_b)
            default_idx_b = list(df_b.columns).index(auto_key_b[0]) if auto_key_b else 0
            st.caption(f"🤖 Auto-detected key column: **{auto_key_b[0] if auto_key_b else 'not found'}**")
            with st.expander("⚙️ Change key column"):
                key_b = st.selectbox("Key column", df_b.columns, index=default_idx_b, key="key_b_select")

    if st.button("Compare Files", type="primary", disabled=not (file_a and file_b)):
        map_a = {normalize(v): idx for idx, v in df_a[key_a].items()}
        map_b = {normalize(v): idx for idx, v in df_b[key_b].items()}
        common_cols = [c for c in df_a.columns if c in df_b.columns]

        changed_rows, only_a_rows, only_b_rows = [], [], []
        for norm_key, idx_a in map_a.items():
            if norm_key not in map_b:
                only_a_rows.append(df_a.loc[idx_a].to_dict())
                continue
            idx_b = map_b[norm_key]
            row_a, row_b = df_a.loc[idx_a], df_b.loc[idx_b]
            for col in common_cols:
                va, vb = str(row_a[col]).strip(), str(row_b[col]).strip()
                if va != vb:
                    changed_rows.append({
                        "Key": row_a[key_a],
                        "Column": col,
                        "Value in POC Project File": va,
                        "Value in Pilot Project File": vb,
                    })
        for norm_key, idx_b in map_b.items():
            if norm_key not in map_a:
                only_b_rows.append(df_b.loc[idx_b].to_dict())

        changed_df = pd.DataFrame(changed_rows)
        if not changed_df.empty:
            vc = changed_df["Column"].value_counts()
            column_summary = pd.DataFrame({"Column": vc.index, "Difference Count": vc.values})
        else:
            column_summary = pd.DataFrame(columns=["Column", "Difference Count"])

        st.session_state.compare_result = {
            "changed": changed_df,
            "only_a": pd.DataFrame(only_a_rows),
            "only_b": pd.DataFrame(only_b_rows),
            "column_summary": column_summary,
        }

    if "compare_result" in st.session_state:
        res = st.session_state.compare_result
        t1, t2, t3, t4 = st.tabs([
            f"Changed ({len(res['changed'])})",
            f"Only in POC File ({len(res['only_a'])})",
            f"Only in Pilot File ({len(res['only_b'])})",
            f"Column Summary ({len(res['column_summary'])})",
        ])
        with t1:
            st.dataframe(res["changed"], use_container_width=True)
        with t2:
            st.dataframe(res["only_a"], use_container_width=True)
        with t3:
            st.dataframe(res["only_b"], use_container_width=True)
        with t4:
            st.dataframe(res["column_summary"], use_container_width=True, hide_index=True)

        st.download_button(
            "⬇ Download Comparison Report",
            data=to_excel_bytes({
                "Changed": res["changed"],
                "Only in POC Project File": res["only_a"],
                "Only in Pilot Project File": res["only_b"],
                "Column-wise Summary": res["column_summary"],
            }),
            file_name="comparison-report.xlsx",
            mime=EXCEL_MIME,
        )

    st.divider()
    st.header("Attribute Completeness Check")
    st.caption(
        "Upload your main project's CGC attribute file — a long-format file with one row per "
        "attribute (e.g. category_name, group_name, class_name, attribute_name, attribute_value). "
        "The tool reads every attribute name found in the file, and for any attribute you select "
        "(e.g. Brand), reports every Category / Group / Class where that attribute is missing or empty."
    )

    if "attribute_options" not in st.session_state:
        st.session_state.attribute_options = load_attribute_options()

    cgc_file = st.file_uploader(
        "Upload the main project's CGC attribute file",
        type=["xlsx", "xls", "csv"], key="cgc_file",
    )

    id_cols, attr_name_col, attr_value_col, df_cgc = [], None, None, None

    if cgc_file:
        df_cgc = read_uploaded_file(cgc_file)
        auto_id_cols = [c for c in df_cgc.columns if CGC_PATTERN.search(c)]
        auto_name_col, auto_value_col = guess_attribute_columns(df_cgc)

        st.success(
            f"🤖 Auto-detected — Identifiers: **{', '.join(auto_id_cols) or 'not found'}** · "
            f"Attribute Name column: **{auto_name_col or 'not found'}** · "
            f"Attribute Value column: **{auto_value_col or 'not found'}**"
        )

        with st.expander("⚙️ Change identifier / attribute columns"):
            id_cols = st.multiselect(
                "Identifier column(s) (Category / Group / Class)",
                list(df_cgc.columns), default=auto_id_cols, key="cgc_id_cols"
            )
            col_options = list(df_cgc.columns)
            name_idx = col_options.index(auto_name_col) if auto_name_col in col_options else 0
            value_idx = col_options.index(auto_value_col) if auto_value_col in col_options else 0
            attr_name_col = st.selectbox("Attribute Name column", col_options, index=name_idx, key="cgc_attr_name_col")
            attr_value_col = st.selectbox("Attribute Value column", col_options, index=value_idx, key="cgc_attr_value_col")

        id_cols = st.session_state.get("cgc_id_cols", auto_id_cols)
        attr_name_col = st.session_state.get("cgc_attr_name_col", auto_name_col)
        attr_value_col = st.session_state.get("cgc_attr_value_col", auto_value_col)

        if attr_name_col:
            discovered = sorted(set(df_cgc[attr_name_col].astype(str).str.strip()) - {""})
            newly_found = [a for a in discovered if a not in st.session_state.attribute_options]
            if newly_found:
                st.session_state.attribute_options.extend(newly_found)
                save_attribute_options(st.session_state.attribute_options)
                st.caption(f"📎 {len(newly_found)} new attribute name(s) discovered in this file and added to the list below.")

    ac1, ac2 = st.columns([4, 1])
    new_attr = ac1.text_input("Add a custom attribute", key="new_attr_input", placeholder="e.g. Brand", label_visibility="collapsed")
    if ac2.button("+ Add Attribute") and new_attr.strip():
        if new_attr.strip() not in st.session_state.attribute_options:
            st.session_state.attribute_options.append(new_attr.strip())
            save_attribute_options(st.session_state.attribute_options)
        st.rerun()
    st.caption("Attribute names discovered in uploaded files, and any added manually, are saved and reused by everyone on this tool.")

    selected_attrs = st.multiselect(
        "Select attributes to check", options=sorted(st.session_state.attribute_options), key="selected_attrs"
    )

    can_check = bool(cgc_file and selected_attrs and id_cols and attr_name_col and attr_value_col)
    if st.button("Check Completeness", type="primary", disabled=not can_check):
        summary_df, missing_df = check_attribute_completeness_long(
            df_cgc, id_cols, attr_name_col, attr_value_col, selected_attrs
        )
        st.session_state.attr_summary = summary_df
        st.session_state.attr_missing = missing_df

    if "attr_summary" in st.session_state:
        st.subheader("Completeness Summary")
        st.dataframe(st.session_state.attr_summary, use_container_width=True, hide_index=True)
        st.subheader("Entries Missing Selected Attributes")
        st.dataframe(st.session_state.attr_missing, use_container_width=True, hide_index=True)
        st.download_button(
            "⬇ Download Attribute Completeness Report",
            data=to_excel_bytes({
                "Summary": st.session_state.attr_summary,
                "Missing Details": st.session_state.attr_missing,
            }),
            file_name="attribute-completeness-report.xlsx",
            mime=EXCEL_MIME,
        )

# =====================================================================
# TOOL 3: System SKU Catalog
# =====================================================================

elif tool == "System SKU Catalog":
    st.header("System SKU Catalog")
    st.caption(
        "Upload your internal SKU/Group/Class reference sheet once — the tool detects the "
        "name and packshot columns automatically. This data is reused in the SKU Match tool."
    )

    catalog_file = st.file_uploader("Upload system catalog file", type=["xlsx", "xls", "csv"], key="catalog_file")
    if catalog_file is not None:
        file_bytes = get_file_bytes(catalog_file)
        st.session_state.system_df = read_df_from_bytes(file_bytes, catalog_file.name)
        st.session_state.system_file_name = catalog_file.name
        st.session_state.system_embedded_images = extract_embedded_images(file_bytes, catalog_file.name)

    if "system_df" in st.session_state:
        df_catalog = st.session_state.system_df
        embedded_map = st.session_state.get("system_embedded_images", {})
        embedded_cols = {c for (_, c) in embedded_map.keys()}

        auto_name_cols = guess_name_columns(df_catalog)
        auto_pack_cols = guess_image_columns(df_catalog, embedded_cols)

        st.success(
            f"🤖 Auto-detected — Name/Group: **{', '.join(auto_name_cols) or 'not found'}** · "
            f"Packshot: **{', '.join(auto_pack_cols) or 'not found'}**"
        )
        if embedded_cols:
            st.caption(f"📎 Embedded images found in: {', '.join(embedded_cols)}")

        with st.expander("⚙️ Advanced: choose columns manually"):
            system_name_cols = st.multiselect(
                "Name/Group/Class column(s)", list(df_catalog.columns), default=auto_name_cols, key="sys_name_cols"
            )
            system_pack_cols = st.multiselect(
                "Packshot column(s)", list(df_catalog.columns), default=auto_pack_cols, key="sys_pack_cols"
            )

        resolved_name_cols = st.session_state.get("sys_name_cols", auto_name_cols)
        resolved_pack_cols = st.session_state.get("sys_pack_cols", auto_pack_cols)

        search = st.text_input("🔍 Search")
        view_df = df_catalog.copy()
        if search:
            mask = view_df.apply(lambda r: search.lower() in " ".join(str(v) for v in r.values).lower(), axis=1)
            view_df = view_df[mask]

        preview_rows = []
        for idx, row in view_df.iterrows():
            pack = get_packshot_value(row, idx, resolved_pack_cols, embedded_map)
            preview_rows.append({**row.to_dict(), "Packshot Preview": pack["display"]})
        display_df = pd.DataFrame(preview_rows) if preview_rows else view_df

        column_config = {}
        if resolved_pack_cols:
            column_config["Packshot Preview"] = st.column_config.ImageColumn("Packshot", width="small")

        st.dataframe(display_df, use_container_width=True, hide_index=True, column_config=column_config)
        st.caption(f"Total {len(df_catalog)} entries in system — {len(view_df)} shown.")
    else:
        st.info("Upload your system catalog file above to get started.")

# =====================================================================
# TOOL 4: SKU Match
# =====================================================================

else:
    st.header("SKU Match")
    st.caption(
        "Match client SKUs against your system catalog regardless of format — image links "
        "and images embedded directly in the sheet are both supported."
    )

    if "system_df" not in st.session_state:
        st.warning("Upload your system sheet in the **'System SKU Catalog'** tab first, then return here to match.")
    else:
        df_internal = st.session_state.system_df
        internal_embedded = st.session_state.get("system_embedded_images", {})
        internal_name_cols = st.session_state.get("sys_name_cols", guess_name_columns(df_internal))
        internal_pack_cols = st.session_state.get(
            "sys_pack_cols", guess_image_columns(df_internal, {c for (_, c) in internal_embedded.keys()})
        )

        st.info(
            f"System catalog: **{st.session_state.get('system_file_name', '')}** "
            f"({len(df_internal)} rows) — loaded from 'System SKU Catalog' tab."
        )

        if not internal_name_cols:
            st.warning("Could not detect a name column in the System SKU Catalog tab.")
        else:
            client_file = st.file_uploader(
                "Upload client SKU file (any format)", type=["xlsx", "xls", "csv"], key="client_file2"
            )

            if client_file:
                client_bytes = get_file_bytes(client_file)
                df_client = read_df_from_bytes(client_bytes, client_file.name)
                client_embedded = extract_embedded_images(client_bytes, client_file.name)
                client_embedded_cols = {c for (_, c) in client_embedded.keys()}

                auto_client_name = guess_name_columns(df_client)
                auto_client_pack = guess_image_columns(df_client, client_embedded_cols)

                st.success(
                    f"🤖 Auto-detected — Name/Group: **{', '.join(auto_client_name) or 'not found'}** · "
                    f"Packshot: **{', '.join(auto_client_pack) or 'not found'}**"
                )
                if client_embedded_cols:
                    st.caption(f"📎 Embedded images found in: {', '.join(client_embedded_cols)}")

                with st.expander("⚙️ Advanced: choose client columns manually"):
                    client_name_cols = st.multiselect(
                        "Client Name/Group column(s)", list(df_client.columns), default=auto_client_name, key="client_name_cols"
                    )
                    client_pack_cols = st.multiselect(
                        "Client Packshot column(s) — select all Front/Back/etc. columns",
                        list(df_client.columns), default=auto_client_pack, key="client_pack_cols",
                    )

                resolved_client_name = st.session_state.get("client_name_cols", auto_client_name)
                resolved_client_pack = st.session_state.get("client_pack_cols", auto_client_pack)

                threshold = st.slider(
                    "Fuzzy match sensitivity", 50, 95, 70, step=5, format="%d%%"
                ) / 100.0
                embed_in_download = st.checkbox(
                    "Embed actual packshot images in the downloaded report (for embedded-source images)",
                    value=True,
                )

                if st.button("🔎 Run Match", type="primary", disabled=not resolved_client_name):
                    internal_records = []
                    for idx, irow in df_internal.iterrows():
                        combined = combine_cols(irow, internal_name_cols)
                        pack = get_packshot_value(irow, idx, internal_pack_cols, internal_embedded)
                        internal_records.append((combined, pack))

                    exact_map = {}
                    for combined, pack in internal_records:
                        norm = normalize(combined)
                        if norm not in exact_map:
                            exact_map[norm] = (combined, pack)

                    results = []
                    client_pack_bytes_map = {}
                    matched_pack_bytes_map = {}

                    for idx, crow in df_client.iterrows():
                        client_combined = combine_cols(crow, resolved_client_name)
                        norm = normalize(client_combined)
                        client_pack = get_packshot_value(crow, idx, resolved_client_pack, client_embedded)

                        match_type, matched_val, confidence = "Unmatched", "", 0.0
                        matched_pack = {"display": "", "bytes": None}

                        if norm in exact_map:
                            matched_val, matched_pack = exact_map[norm]
                            match_type, confidence = "Exact", 1.0
                        else:
                            best_score, best_combined, best_pack = 0.0, None, {"display": "", "bytes": None}
                            for combined, pack in internal_records:
                                sc = similarity(client_combined, combined)
                                if sc > best_score:
                                    best_score, best_combined, best_pack = sc, combined, pack
                            if best_combined is not None and best_score >= threshold:
                                matched_val, match_type, confidence, matched_pack = best_combined, "Fuzzy", best_score, best_pack

                        row_num = len(results)
                        if client_pack["bytes"]:
                            client_pack_bytes_map[row_num] = client_pack["bytes"]
                        if matched_pack["bytes"]:
                            matched_pack_bytes_map[row_num] = matched_pack["bytes"]

                        results.append({
                            "Client SKU/Group": client_combined,
                            "Client Packshot": client_pack["display"],
                            "Matched SKU (System)": matched_val,
                            "Matched Packshot": matched_pack["display"],
                            "Match Type": match_type,
                            "Confidence %": round(confidence * 100),
                        })

                    st.session_state.match_results = pd.DataFrame(results)
                    st.session_state.match_client_pack_bytes = client_pack_bytes_map
                    st.session_state.match_matched_pack_bytes = matched_pack_bytes_map
                    st.session_state.match_embed_flag = embed_in_download

                if "match_results" in st.session_state:
                    results_df = st.session_state.match_results
                    total = len(results_df)
                    matched_count = int((results_df["Match Type"] != "Unmatched").sum()) if total else 0
                    match_rate = round(matched_count / total * 100) if total else 0
                    counts = results_df["Match Type"].value_counts().to_dict()

                    m0, m1, m2, m3 = st.columns(4)
                    m0.metric("Match Rate", f"{match_rate}%")
                    m1.metric("Exact", counts.get("Exact", 0))
                    m2.metric("Fuzzy", counts.get("Fuzzy", 0))
                    m3.metric("Unmatched", counts.get("Unmatched", 0))

                    view = st.radio("View", ["All", "Fuzzy", "Unmatched"], horizontal=True)
                    if view == "Fuzzy":
                        show_df = results_df[results_df["Match Type"] == "Fuzzy"]
                    elif view == "Unmatched":
                        show_df = results_df[results_df["Match Type"] == "Unmatched"]
                    else:
                        show_df = results_df

                    st.dataframe(
                        show_df,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Client Packshot": st.column_config.ImageColumn("Client Packshot", width="small"),
                            "Matched Packshot": st.column_config.ImageColumn("Matched Packshot", width="small"),
                            "Confidence %": st.column_config.ProgressColumn(
                                "Confidence", min_value=0, max_value=100, format="%d%%"
                            ),
                        },
                    )

                    embed_flag = st.session_state.get("match_embed_flag", True)
                    excel_bytes = build_match_excel(
                        results_df,
                        {
                            "client": st.session_state.get("match_client_pack_bytes", {}) if embed_flag else {},
                            "matched": st.session_state.get("match_matched_pack_bytes", {}) if embed_flag else {},
                        },
                    )
                    st.download_button(
                        "⬇ Download Match Report",
                        data=excel_bytes,
                        file_name="sku-match-report.xlsx",
                        mime=EXCEL_MIME,
                    )
